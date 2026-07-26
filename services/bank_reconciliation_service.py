import datetime, csv, os
from decimal import Decimal
from typing import List, Optional
from sqlalchemy.orm import Session
from sqlalchemy import func
from models import (
    BankStatement, BankTransaction, TreasuryTransaction,
    JournalEntry, JournalLine, DECIMAL_ZERO,
)
from currency_utils import D
from services.accounting_service import AccountingService


class BankReconciliationService:
    def __init__(self, db: Session):
        self.db = db

    def import_statement(self, account_id: int, file_path: str,
                         period_start: str = None, period_end: str = None,
                         created_by: str = None) -> BankStatement:
        statement = BankStatement(
            account_id=account_id,
            period_start=datetime.date.fromisoformat(period_start) if period_start else None,
            period_end=datetime.date.fromisoformat(period_end) if period_end else None,
            opening_balance=DECIMAL_ZERO,
            closing_balance=DECIMAL_ZERO,
            file_path=file_path,
            status="imported",
            created_by=created_by,
        )
        self.db.add(statement)
        self.db.flush()

        rows = self._read_file(file_path)
        for row in rows:
            date = self._parse_date(row.get("date"))
            debit = D(str(row.get("debit", 0) or 0))
            credit = D(str(row.get("credit", 0) or 0))
            bal = D(str(row.get("balance", 0))) if row.get("balance") else None

            txn = BankTransaction(
                statement_id=statement.id,
                date=date,
                description=str(row.get("description", row.get("details", row.get("narration", "")))),
                reference=str(row.get("reference", row.get("ref", row.get("check", "")))),
                debit=debit,
                credit=credit,
                balance=bal,
                match_status="unmatched",
            )
            self.db.add(txn)

        if rows:
            first = rows[0]
            last = rows[-1]
            if first.get("balance"):
                statement.opening_balance = D(str(first["balance"]))
            if last.get("balance"):
                statement.closing_balance = D(str(last["balance"]))

        self.db.flush()
        self._auto_match(statement.id)
        return statement

    def _read_file(self, file_path: str) -> list:
        ext = file_path.rsplit(".", 1)[-1].lower()
        if ext == "csv":
            return self._read_csv(file_path)
        elif ext in ("xls", "xlsx"):
            return self._read_excel(file_path)
        else:
            raise ValueError(f"Unsupported file type: {ext}")

    def _read_csv(self, file_path: str) -> list:
        path = file_path.lstrip(".")
        if not os.path.isabs(path):
            path = os.path.join(os.getcwd(), path)
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.DictReader(f)
            rows = []
            for r in reader:
                rows.append({k.strip().lower(): v.strip() if v else "" for k, v in r.items()})
        return rows

    def _read_excel(self, file_path: str) -> list:
        from openpyxl import load_workbook
        path = file_path.lstrip(".")
        if not os.path.isabs(path):
            path = os.path.join(os.getcwd(), path)
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1))
        headers = [c.value.strip().lower() if c.value else "" for c in header_row]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_data = {}
            for i, val in enumerate(row):
                h = headers[i] if i < len(headers) else f"col{i}"
                row_data[h] = val if val is not None else ""
            rows.append(row_data)
        wb.close()
        return rows

    def _parse_date(self, val) -> datetime.date:
        if isinstance(val, datetime.datetime):
            return val.date()
        if isinstance(val, datetime.date):
            return val
        try:
            return datetime.datetime.strptime(str(val)[:10], "%Y-%m-%d").date()
        except ValueError:
            pass
        try:
            return datetime.datetime.strptime(str(val)[:10], "%d/%m/%Y").date()
        except ValueError:
            pass
        try:
            return datetime.datetime.strptime(str(val)[:10], "%m/%d/%Y").date()
        except ValueError:
            pass
        return datetime.date.today()

    def _auto_match(self, statement_id: int):
        """مطابقة تلقائية: reference + amount + date"""
        stmt = self.db.query(BankStatement).get(statement_id)
        if not stmt:
            return
        account_id = stmt.account_id
        bank_txns = self.db.query(BankTransaction).filter(
            BankTransaction.statement_id == statement_id,
            BankTransaction.match_status == "unmatched",
        ).all()

        for bt in bank_txns:
            matched = self._find_match(bt, account_id)
            if matched:
                bt.matched_txn_id = matched.id
                bt.match_status = "matched"
                self.db.flush()

        self._mark_duplicates(statement_id)
        self.db.flush()

    def _find_match(self, bt: BankTransaction, account_id: int) -> Optional[TreasuryTransaction]:
        """البحث عن TreasuryTransaction مطابقة"""
        q = self.db.query(TreasuryTransaction).filter(
            TreasuryTransaction.account_id == account_id,
            TreasuryTransaction.date == bt.date,
        )
        if bt.reference and bt.reference.strip() and bt.reference != "nan":
            q_ref = q.filter(TreasuryTransaction.reference == bt.reference.strip())
            txn = q_ref.first()
            if txn:
                return txn

        amount = bt.debit if bt.debit > 0 else bt.credit
        q_amt = q.filter(TreasuryTransaction.amount == amount)
        txn = q_amt.first()
        if txn:
            return txn

        q_desc = q.filter(
            func.lower(TreasuryTransaction.description).contains(
                bt.description.lower()[:30]
            ) if bt.description else False
        ).first()
        if q_desc:
            return q_desc

        return None

    def _mark_duplicates(self, statement_id: int):
        """تحديد المكرر داخل نفس الكشف"""
        refs = self.db.query(
            BankTransaction.reference, func.count(BankTransaction.id)
        ).filter(
            BankTransaction.statement_id == statement_id,
            BankTransaction.match_status == "unmatched",
            BankTransaction.reference.isnot(None),
            BankTransaction.reference != "",
            BankTransaction.reference != "nan",
        ).group_by(BankTransaction.reference).having(func.count(BankTransaction.id) > 1).all()

        for ref, cnt in refs:
            txns = self.db.query(BankTransaction).filter(
                BankTransaction.statement_id == statement_id,
                BankTransaction.reference == ref,
                BankTransaction.match_status == "unmatched",
            ).order_by(BankTransaction.id).all()
            for t in txns[1:]:
                t.match_status = "duplicate"

    def get_unmatched(self, statement_id: int) -> List[BankTransaction]:
        return self.db.query(BankTransaction).filter(
            BankTransaction.statement_id == statement_id,
            BankTransaction.match_status.in_(["unmatched", "duplicate"]),
        ).all()

    def get_matched(self, statement_id: int) -> List[BankTransaction]:
        return self.db.query(BankTransaction).filter(
            BankTransaction.statement_id == statement_id,
            BankTransaction.match_status == "matched",
        ).all()

    def manual_match(self, bank_txn_id: int, txn_id: int):
        bt = self.db.query(BankTransaction).get(bank_txn_id)
        if bt:
            bt.matched_txn_id = txn_id
            bt.match_status = "matched"
            self.db.flush()

    def manual_unmatch(self, bank_txn_id: int):
        bt = self.db.query(BankTransaction).get(bank_txn_id)
        if bt:
            bt.matched_txn_id = None
            bt.match_status = "unmatched"
            self.db.flush()

    def mark_bank_fee(self, bank_txn_id: int, notes: str = ""):
        """تحديد حركة بنكية كرسوم بنكية"""
        bt = self.db.query(BankTransaction).get(bank_txn_id)
        if bt:
            bt.match_status = "bank_fee"
            bt.notes = notes
            self.db.flush()

    def create_fee_entry(self, bank_txn_id: int, created_by: str = None):
        """إنشاء قيد يومي للرسوم البنكية"""
        bt = self.db.query(BankTransaction).get(bank_txn_id)
        if not bt or bt.match_status != "bank_fee":
            return
        asc = AccountingService(self.db)
        treasury_acc = asc._acc_id("treasury")
        opex_acc = asc._acc_id("opex")
        from models import generate_journal_entry_number
        entry = JournalEntry(
            number=generate_journal_entry_number(self.db),
            date=bt.date,
            description=f"رسوم بنكية - {bt.description or ''}",
            source_type="bank_reconciliation",
            source_id=bt.id,
            created_by=created_by,
            status="posted",
        )
        self.db.add(entry)
        self.db.flush()
        self.db.add_all([
            JournalLine(journal_entry_id=entry.id, account_id=opex_acc, debit=bt.debit, credit=DECIMAL_ZERO),
            JournalLine(journal_entry_id=entry.id, account_id=treasury_acc, credit=bt.debit, debit=DECIMAL_ZERO),
        ])
        self.db.flush()

    def get_statement_summary(self, statement_id: int) -> dict:
        stmt = self.db.query(BankStatement).get(statement_id)
        if not stmt:
            return {}
        total_bank = self.db.query(
            func.sum(BankTransaction.debit), func.sum(BankTransaction.credit)
        ).filter(BankTransaction.statement_id == statement_id).first()
        total_internal = self.db.query(
            func.sum(TreasuryTransaction.amount)
        ).filter(
            TreasuryTransaction.account_id == stmt.account_id,
            TreasuryTransaction.date.between(stmt.period_start, stmt.period_end) if stmt.period_start and stmt.period_end else True,
        ).first()
        unmatched = self.db.query(BankTransaction).filter(
            BankTransaction.statement_id == statement_id,
            BankTransaction.match_status == "unmatched",
        ).count()
        matched = self.db.query(BankTransaction).filter(
            BankTransaction.statement_id == statement_id,
            BankTransaction.match_status == "matched",
        ).count()
        duplicates = self.db.query(BankTransaction).filter(
            BankTransaction.statement_id == statement_id,
            BankTransaction.match_status == "duplicate",
        ).count()
        fees = self.db.query(BankTransaction).filter(
            BankTransaction.statement_id == statement_id,
            BankTransaction.match_status == "bank_fee",
        ).count()
        return {
            "total_debit": D(str(total_bank[0])) if total_bank and total_bank[0] else DECIMAL_ZERO,
            "total_credit": D(str(total_bank[1])) if total_bank and total_bank[1] else DECIMAL_ZERO,
            "unmatched": unmatched,
            "matched": matched,
            "duplicates": duplicates,
            "bank_fees": fees,
        }
