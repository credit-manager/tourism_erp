import io, datetime
from fastapi import Request, Depends
from fastapi.responses import HTMLResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from currency_utils import D, DECIMAL_ZERO

from . import templates, get_db, parse_report_dates, build_report_data
from models import Customer, Reservation
import auth


def setup_report_routes(app):
    @app.get("/reports", response_class=HTMLResponse)
    def reports_page(request: Request, date_from: str = None, date_to: str = None,
                     db: Session = Depends(get_db), user=Depends(auth.require_permission("reports.view"))):
        d_from, d_to = parse_report_dates(date_from, date_to)
        data = build_report_data(db, d_from, d_to)
        return templates.TemplateResponse(request, "reports.html", {
            "request": request, "page_title": "التقارير - الربح والخسارة", "active": "reports",
            "date_from": d_from.isoformat(), "date_to": d_to.isoformat(), **data,
        })

    @app.get("/reports/export/excel")
    def export_report_excel(date_from: str = None, date_to: str = None, db: Session = Depends(get_db),
                            user=Depends(auth.require_permission("reports.view"))):
        d_from, d_to = parse_report_dates(date_from, date_to)
        data = build_report_data(db, d_from, d_to)
        wb = Workbook()
        ws = wb.active
        ws.title = "الملخص"
        header_fill = PatternFill(start_color="6E45A3", end_color="6E45A3", fill_type="solid")
        ws.append(["تقرير الربح والخسارة"])
        ws["A1"].font = Font(bold=True, size=14)
        ws.append([f"من {d_from.isoformat()} إلى {d_to.isoformat()}"])
        ws.append([])
        rows = [
            ("إجمالي المبيعات (على العميل)", data["total_revenue"]),
            ("إجمالي تكلفة الإقامة للفنادق", data["total_hotel_cost"]),
            ("الربح الإجمالي قبل العمولات والمصروفات", data["gross_profit"]),
            ("إجمالي عمولات الموظفين", data["total_commissions"]),
            ("إجمالي المصروفات التشغيلية", data["total_expenses"]),
            ("صافي الربح", data["net_profit"]),
        ]
        for label, value in rows:
            ws.append([label, round(value, 2)])
        for row in ws.iter_rows(min_row=4, max_row=9):
            row[0].font = Font(bold=True)
        ws2 = wb.create_sheet("الحجوزات")
        headers = ["رقم الحجز", "التاريخ", "النزيل", "الفندق", "تكلفة الشركة", "تكلفة الإقامة", "ربح الحجز", "عمولة الموظف"]
        ws2.append(headers)
        for cell in ws2[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        for r in data["reservations"]:
            ws2.append([r.booking_number, r.created_date.isoformat() if r.created_date else "",
                        r.guest_name, r.hotel.name if r.hotel else "", r.company_cost or 0,
                        r.stay_cost or 0, r.profit, r.employee_commission or 0])
        ws3 = wb.create_sheet("المصروفات")
        ws3.append(["التاريخ", "التصنيف", "المبلغ", "الوصف"])
        for cell in ws3[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        for e in data["expenses"]:
            ws3.append([e.date.isoformat() if e.date else "", e.category, e.amount or 0, e.description])
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
                sheet.column_dimensions[col[0].column_letter].width = max(12, min(35, max_len + 3))
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"profit_loss_report_{d_from}_{d_to}.xlsx"
        return StreamingResponse(
            buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    @app.get("/reports/customer-credit", response_class=HTMLResponse)
    def customer_credit_report(request: Request, db: Session = Depends(get_db),
                                user=Depends(auth.require_permission("reports.view"))):
        """Report of customers with credit limit issues."""
        customers = db.query(Customer).filter(
            Customer.credit_limit > DECIMAL_ZERO,
            or_(Customer.type == "B2B", Customer.type == "B2C")
        ).all()

        rows = []
        for c in customers:
            open_res = db.query(Reservation).filter(
                Reservation.customer_id == c.id,
                Reservation.status.notin_(["cancelled", "closed"]),
            ).all()
            outstanding = sum(max(DECIMAL_ZERO, r.remaining_to_office) for r in open_res)
            open_count = len(open_res)
            available = D(c.credit_limit) - outstanding
            exceeded = outstanding > D(c.credit_limit)
            # Overdue: credit reservations past due date
            overdue = db.query(Reservation).filter(
                Reservation.customer_id == c.id,
                Reservation.reservation_type == "credit",
                Reservation.payment_due_date != None,
                Reservation.payment_due_date < datetime.date.today(),
                Reservation.status.notin_(["cancelled", "closed"]),
            ).all()
            overdue_total = sum(max(DECIMAL_ZERO, r.remaining_to_office) for r in overdue)
            overdue_count = len(overdue)
            payment_terms_label = {
                "due_on_receipt": "عند الاستلام", "net_15": "15 يوم",
                "net_30": "30 يوم", "net_60": "60 يوم", "custom": "مخصص"
            }.get(c.payment_terms or "", "—")
            rows.append({
                "customer": c,
                "credit_limit": c.credit_limit,
                "outstanding": outstanding,
                "open_count": open_count,
                "available": max(DECIMAL_ZERO, available),
                "exceeded": exceeded,
                "overdue_total": overdue_total,
                "overdue_count": overdue_count,
                "payment_terms_label": payment_terms_label,
            })
        return templates.TemplateResponse(request, "report_customer_credit.html", {
            "request": request, "page_title": "تقرير الائتمان", "active": "reports",
            "rows": rows,
        })
