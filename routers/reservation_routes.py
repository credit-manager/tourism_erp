import datetime, io
from urllib.parse import quote
from fastapi import Request, Form, Depends, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from currency_utils import D, DECIMAL_ZERO

from models import (
    Reservation, Hotel, Employee, Customer, Service, Supplier, generate_booking_number,
    Collection, CollectionAllocation, generate_collection_number,
    TreasuryAccount, TreasuryTransaction, JournalEntry,
    ReservationDocument, STATUS_LABELS, STATUS_LABELS_EN, STATUS_COLORS,
    RoomType, MealPlan, SecureFile,
)
from . import (
    templates, get_db, _parse_date, _apply_reservation_fields,
    _detect_conflicts,
    _find_hotel_by_exact_name, _get_or_create_hotel, _get_or_create_employee,
    _parse_excel_date, RESERVATION_EXCEL_COLUMNS, save_upload,
    paginate, pagination_ctx,
)
import auth
from services.storage_service import storage_service, StorageError, StorageSecurityError


def setup_reservation_routes(app):
    @app.get("/reservations/board", response_class=HTMLResponse)
    def reservations_board(request: Request, selected: int = None, db: Session = Depends(get_db),
                           user=Depends(auth.require_permission("reservations.view"))):
        reservations = db.query(Reservation).order_by(Reservation.id.desc()).limit(100).all()
        active = None
        if selected:
            active = db.query(Reservation).get(selected)
        elif reservations:
            active = reservations[0]
        return templates.TemplateResponse(request, "reservations_board.html", {
            "request": request, "page_title": "لوحة الحجوزات", "active": "board",
            "reservations": reservations, "active_res": active,
        })

    @app.get("/reservations", response_class=HTMLResponse)
    def reservations_page(request: Request, db: Session = Depends(get_db),
                          q: str = None, status_filter: str = "open",
                          booking_date_from: str = None, booking_date_to: str = None,
                          checkin_from: str = None, checkin_to: str = None,
                          checkout_from: str = None, checkout_to: str = None,
                          hotel_id: str = None, page: int = 1):
        query = db.query(Reservation)
        if hotel_id:
            try: query = query.filter(Reservation.hotel_id == int(hotel_id))
            except Exception: pass
        if booking_date_from:
            d = _parse_date(booking_date_from)
            if d: query = query.filter(Reservation.created_date >= d)
        if booking_date_to:
            d = _parse_date(booking_date_to)
            if d: query = query.filter(Reservation.created_date <= d)
        if checkin_from:
            d = _parse_date(checkin_from)
            if d: query = query.filter(Reservation.checkin_date >= d)
        if checkin_to:
            d = _parse_date(checkin_to)
            if d: query = query.filter(Reservation.checkin_date <= d)
        if checkout_from:
            d = _parse_date(checkout_from)
            if d: query = query.filter(Reservation.checkout_date >= d)
        if checkout_to:
            d = _parse_date(checkout_to)
            if d: query = query.filter(Reservation.checkout_date <= d)
        all_matching = query.order_by(Reservation.id.desc()).all()
        if q:
            ql = q.lower().strip()
            def matches(r):
                haystack = " ".join(str(x or "") for x in [
                    r.booking_number, r.created_date, r.guest_name, r.phone, r.passport_no,
                    r.nationality, r.room_type, r.stay_type, r.notes,
                    r.hotel.name if r.hotel else "", r.employee.name if r.employee else "",
                ]).lower()
                return ql in haystack
            all_matching = [r for r in all_matching if matches(r)]
        if status_filter == "open":
            reservations = [r for r in all_matching if not r.is_paid_in_full]
        elif status_filter == "paid":
            reservations = [r for r in all_matching if r.is_paid_in_full]
        elif status_filter == "cancelled":
            reservations = [r for r in all_matching if r.status == "cancelled"]
        else:
            reservations = [r for r in all_matching if r.status != "cancelled"]
        total_all = len(reservations)
        per_page = 30
        total_pages = max(1, (total_all + per_page - 1) // per_page)
        page = max(1, min(page, total_pages))
        start = (page - 1) * per_page
        reservations_page = reservations[start:start + per_page]
        hotels = db.query(Hotel).all()
        employees = db.query(Employee).all()
        customers = db.query(Customer).all()
        services = db.query(Service).all()
        suppliers = db.query(Supplier).all()
        total_amount = sum(D(r.company_cost) for r in reservations_page)
        total_remaining = sum(r.remaining_to_office for r in reservations_page)
        all_res = db.query(Reservation).all()
        open_count = sum(1 for r in all_res if not r.is_paid_in_full)
        pending_count = sum(1 for r in all_res if r.status == "pending")
        base_url = f"/reservations?page={page}"
        if q: base_url += f"&q={q}"
        if status_filter: base_url += f"&status_filter={status_filter}"
        if booking_date_from: base_url += f"&booking_date_from={booking_date_from}"
        if booking_date_to: base_url += f"&booking_date_to={booking_date_to}"
        if checkin_from: base_url += f"&checkin_from={checkin_from}"
        if checkin_to: base_url += f"&checkin_to={checkin_to}"
        if checkout_from: base_url += f"&checkout_from={checkout_from}"
        if checkout_to: base_url += f"&checkout_to={checkout_to}"
        if hotel_id: base_url += f"&hotel_id={hotel_id}"
        p_ctx = {"page": page, "total_pages": total_pages, "total": total_all,
                 "has_prev": page > 1, "has_next": page < total_pages,
                 "prev_page": page - 1, "next_page": page + 1}
        room_types_all = db.query(RoomType).all()
        meal_plans_all = db.query(MealPlan).all()
        return templates.TemplateResponse(request, "reservations.html", {
            "request": request, "page_title": "الحجوزات (Reservations)", "active": "reservations",
            "reservations": reservations_page, "hotels": hotels, "employees": employees,
            "customers": customers, "services": services, "suppliers": suppliers,
            "total_amount": total_amount, "total_remaining": total_remaining,
            "open_count": open_count, "pending_count": pending_count,
            "q": q or "", "status_filter": status_filter or "open",
            "booking_date_from": booking_date_from or "", "booking_date_to": booking_date_to or "",
            "checkin_from": checkin_from or "", "checkin_to": checkin_to or "",
            "checkout_from": checkout_from or "", "checkout_to": checkout_to or "",
            "selected_hotel_id": hotel_id or "",
            "STATUS_LABELS": STATUS_LABELS, "STATUS_LABELS_EN": STATUS_LABELS_EN,
            "STATUS_COLORS": STATUS_COLORS,
            "conflict_msg": request.query_params.get("conflict_msg"),
            "p": p_ctx, "base_url": base_url,
            "room_types": room_types_all, "meal_plans": meal_plans_all,
        })

    @app.post("/reservations/add")
    async def add_reservation(request: Request, db: Session = Depends(get_db),
                              user=Depends(auth.require_permission("reservations.add"))):
        from services.reservation_service import ReservationService
        form = await request.form()
        try:
            result = ReservationService(db, user).create_reservation(form)
        except Exception as e:
            db.rollback()
            return RedirectResponse(f"/reservations?error=posting_failed&msg={e}", status_code=303)
        if "conflict" in result:
            numbers = result["numbers"]
            return RedirectResponse(
                f"/reservations?conflict=1&conflict_msg=يوجد+حجز+آخر+لنفس+العميل+في+نفس+الفندق+بتواريخ+متداخلة:+{numbers}",
                status_code=303)
        if "error" in result:
            return RedirectResponse(f"/reservations?error=duplicate_number&msg=رقم+الحجز+{result['number']}+مستخدم+بالفعل", status_code=303)
        if "credit_limit" in result:
            return RedirectResponse(f"/reservations?error=credit_limit&msg={result['msg']}", status_code=303)
        reservation = result["reservation"]
        db.commit()
        return RedirectResponse("/reservations", status_code=303)

    @app.post("/reservations/{reservation_id}/confirm")
    def confirm_reservation(reservation_id: int, override_credit: bool = Form(False),
                            db: Session = Depends(get_db),
                            user=Depends(auth.require_permission("reservations.confirm"))):
        from services.reservation_service import ReservationService
        try:
            result = ReservationService(db, user).confirm(reservation_id)
        except Exception as e:
            db.rollback()
            return RedirectResponse(f"/reservations?error=posting_failed&msg={e}", status_code=303)
        if isinstance(result, dict) and result.get("credit_limit"):
            return RedirectResponse(f"/reservations?error=credit_limit&msg={result['msg']}", status_code=303)
        db.commit()
        return RedirectResponse("/reservations", status_code=303)

    @app.get("/reservations/{reservation_id}/edit", response_class=HTMLResponse)
    def edit_reservation_page(reservation_id: int, request: Request, db: Session = Depends(get_db),
                              user=Depends(auth.require_permission("reservations.view"))):
        reservation = db.query(Reservation).get(reservation_id)
        hotels = db.query(Hotel).all()
        employees = db.query(Employee).all()
        customers = db.query(Customer).all()
        services = db.query(Service).all()
        suppliers = db.query(Supplier).all()
        selected_service_ids = [s.id for s in reservation.services] if reservation else []
        from models import STATUS_LABELS as _SL, STATUS_LABELS_EN as _SL_EN
        from state_machine import allowed_transitions, STATE_LABELS_AR, STATE_LABELS_EN as ST_EN, STATE_COLORS
        from models import StateLog
        from timeline import build_reservation_timeline
        def get_logs(rid):
            return db.query(StateLog).filter(StateLog.reservation_id == rid).order_by(StateLog.timestamp.desc()).all()
        timeline_items = build_reservation_timeline(db, reservation_id) if reservation else []
        room_types_all = db.query(RoomType).filter(RoomType.hotel_id == reservation.hotel_id).all() if reservation else []
        meal_plans_all = db.query(MealPlan).filter(MealPlan.hotel_id == reservation.hotel_id).all() if reservation else []
        return templates.TemplateResponse(request, "reservation_edit.html", {
            "request": request, "page_title": f"تعديل الحجز {reservation.booking_number}" if reservation else "تعديل الحجز",
            "active": "reservations", "r": reservation, "hotels": hotels, "employees": employees,
            "customers": customers, "services": services, "suppliers": suppliers,
            "selected_service_ids": selected_service_ids,
            "STATUS_LABELS": _SL, "STATUS_LABELS_EN": _SL_EN,
            "allowed_transitions": allowed_transitions,
            "state_logs": get_logs, "STATE_LABELS_AR": STATE_LABELS_AR,
            "STATE_LABELS_EN": ST_EN, "STATE_COLORS": STATE_COLORS,
            "timeline_items": timeline_items,
            "room_types": room_types_all, "meal_plans": meal_plans_all,
        })

    @app.post("/reservations/{reservation_id}/edit")
    async def update_reservation(reservation_id: int, request: Request, db: Session = Depends(get_db),
                                 user=Depends(auth.require_permission("reservations.edit"))):
        reservation = db.query(Reservation).get(reservation_id)
        if not reservation:
            return RedirectResponse("/reservations", status_code=303)
        form = await request.form()
        data = dict(form)
        data["service_ids"] = [int(x) for x in form.getlist("service_ids")] if form.getlist("service_ids") else []
        new_num = str(data.get("new_booking_number") or "").strip()
        if new_num and new_num != reservation.booking_number:
            conflict = db.query(Reservation).filter(Reservation.booking_number == new_num, Reservation.id != reservation_id).first()
            if conflict:
                return RedirectResponse(f"/reservations/{reservation_id}/edit?error=duplicate&msg=رقم+{new_num}+مستخدم+بالفعل", status_code=303)
            reservation.booking_number = new_num
        from services.reservation_service import ReservationService
        from services.commission_service import CommissionService
        svc = ReservationService(db, user)
        svc.modify(reservation.id, data, reason=data.get("modification_reason", ""))
        CommissionService(db).apply_policies(reservation)
        # Auto-create workflow tasks if updated to confirmed
        if reservation.status == "confirmed":
            from workflow_mappings import get_tasks_for_reservation
            from models import WorkflowTask
            import datetime as _dt
            task_defs = get_tasks_for_reservation(reservation, db)
            existing_titles = set()
            for wt in db.query(WorkflowTask).filter(WorkflowTask.reservation_id == reservation.id).all():
                existing_titles.add(wt.title)
            for td, svc_name in task_defs:
                if td.title_ar in existing_titles or td.title_en in existing_titles:
                    continue
                due = None
                if td.due_days_from_now is not None:
                    due = _dt.date.today() + _dt.timedelta(days=td.due_days_from_now)
                elif reservation.checkin_date and td.due_days_from_checkin:
                    due = reservation.checkin_date + _dt.timedelta(days=td.due_days_from_checkin)
                db.add(WorkflowTask(
                    reservation_id=reservation.id,
                    title=td.title_ar,
                    description=f"{td.title_en} | {svc_name}" if svc_name else td.title_en,
                    priority=td.priority,
                    reminder=td.reminder,
                    due_date=due,
                ))
        db.commit()
        return RedirectResponse("/reservations", status_code=303)

    @app.post("/reservations/{reservation_id}/quick-pay")
    def quick_pay_reservation(reservation_id: int, amount: float = Form(...), payment_method: str = Form("cash"),
                              notes: str = Form(""), db: Session = Depends(get_db),
                              user=Depends(auth.require_permission("reservations.edit"))):
        from services.reservation_service import ReservationService
        try:
            ReservationService(db, user).quick_pay(reservation_id, D(amount), payment_method, notes)
            db.commit()
        except Exception as e:
            print(f"Quick pay failed: {e}")
        return RedirectResponse("/reservations", status_code=303)

    @app.post("/reservations/{reservation_id}/duplicate")
    def duplicate_reservation(reservation_id: int, db: Session = Depends(get_db),
                              user=Depends(auth.require_permission("reservations.add"))):
        from services.reservation_service import ReservationService
        clone = ReservationService(db, user).duplicate(reservation_id)
        db.commit()
        return RedirectResponse(f"/reservations/{clone.id}/edit", status_code=303)

    @app.post("/reservations/{reservation_id}/documents/upload")
    async def upload_reservation_document(reservation_id: int, doc_type: str = Form("other"),
                                          file: UploadFile = File(...), request: Request = None,
                                          db: Session = Depends(get_db),
                                          user=Depends(auth.require_permission("reservations.edit"))):
        reservation = db.query(Reservation).get(reservation_id)
        if reservation and file and file.filename:
            try:
                sf = storage_service.save(file, owner_user_id=getattr(user, "id", None),
                                          category=doc_type or "doc", record_type="reservation",
                                          record_id=reservation_id, db=db, request=request)
                db.add(ReservationDocument(reservation_id=reservation_id, doc_type=doc_type,
                                           file_path=f"/secure-files/{sf.id}",
                                           original_name=file.filename,
                                           uploaded_at=datetime.date.today()))
            except (StorageError, StorageSecurityError):
                return RedirectResponse(f"/reservations/{reservation_id}/edit?error=upload", status_code=303)
            db.commit()
        return RedirectResponse(f"/reservations/{reservation_id}/edit", status_code=303)

    @app.post("/reservations/documents/{document_id}/delete")
    def delete_reservation_document(document_id: int, request: Request = None, db: Session = Depends(get_db),
                                    user=Depends(auth.require_permission("reservations.edit"))):
        doc = db.query(ReservationDocument).get(document_id)
        if doc:
            rid = doc.reservation_id
            if doc.file_path and doc.file_path.startswith("/secure-files/"):
                try:
                    sf = db.query(SecureFile).get(int(doc.file_path.rsplit("/", 1)[-1]))
                    if sf:
                        storage_service.delete(sf, db=db, request=request)
                        db.delete(sf)
                except Exception:
                    pass
            db.delete(doc)
            db.commit()
            return RedirectResponse(f"/reservations/{rid}/edit", status_code=303)
        return RedirectResponse("/reservations", status_code=303)

    @app.get("/reservations/{reservation_id}/voucher", response_class=HTMLResponse)
    def reservation_voucher(reservation_id: int, request: Request, db: Session = Depends(get_db),
                            user=Depends(auth.require_permission("reservations.view"))):
        reservation = db.query(Reservation).get(reservation_id)
        if not reservation:
            return RedirectResponse("/reservations", status_code=303)
        return templates.TemplateResponse(request, "reservation_voucher.html", {"request": request, "r": reservation})

    @app.get("/reservations/{reservation_id}/send-confirmation", response_class=HTMLResponse)
    def send_confirmation(reservation_id: int, request: Request, db: Session = Depends(get_db),
                          user=Depends(auth.require_permission("reservations.view"))):
        reservation = db.query(Reservation).get(reservation_id)
        if not reservation:
            return RedirectResponse("/reservations", status_code=303)
        message = (
            f"تأكيد حجز رقم {reservation.booking_number}\\n"
            f"العميل: {(reservation.customer.name if reservation.customer else reservation.guest_name)}\\n"
            f"الفندق: {reservation.hotel.name if reservation.hotel else '-'}\\n"
            f"الوصول: {reservation.checkin_date} | المغادرة: {reservation.checkout_date}\\n"
            f"عدد الليالي: {reservation.nights} | الغرف: {reservation.room_count}\\n"
            f"إجمالي المبلغ: {reservation.net_sale_price:.2f}"
        )
        return templates.TemplateResponse(request, "send_confirmation.html", {
            "request": request, "r": reservation, "message": message,
        })

    @app.get("/reservations/export/excel")
    def export_reservations_excel(request: Request, db: Session = Depends(get_db),
                                  q: str = None, status_filter: str = "all",
                                  checkin_from: str = None, checkin_to: str = None,
                                  checkout_from: str = None, checkout_to: str = None,
                                  hotel_id: int = None,
                                  user=Depends(auth.require_permission("reservations.view"))):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        query = db.query(Reservation)
        if hotel_id:
            query = query.filter(Reservation.hotel_id == hotel_id)
        if checkin_from:
            d = _parse_date(checkin_from)
            if d: query = query.filter(Reservation.checkin_date >= d)
        if checkin_to:
            d = _parse_date(checkin_to)
            if d: query = query.filter(Reservation.checkin_date <= d)
        if checkout_from:
            d = _parse_date(checkout_from)
            if d: query = query.filter(Reservation.checkout_date >= d)
        if checkout_to:
            d = _parse_date(checkout_to)
            if d: query = query.filter(Reservation.checkout_date <= d)
        all_matching = query.order_by(Reservation.id.desc()).all()
        if q:
            ql = q.lower().strip()
            def matches(r):
                haystack = " ".join(str(x or "") for x in [
                    r.booking_number, r.guest_name, r.phone, r.passport_no, r.nationality,
                    r.room_type, r.stay_type, r.notes, r.hotel.name if r.hotel else "",
                    r.employee.name if r.employee else "",
                ]).lower()
                return ql in haystack
            all_matching = [r for r in all_matching if matches(r)]
        if status_filter == "open":
            reservations = [r for r in all_matching if not r.is_paid_in_full]
        elif status_filter == "paid":
            reservations = [r for r in all_matching if r.is_paid_in_full]
        elif status_filter == "cancelled":
            reservations = [r for r in all_matching if r.status == "cancelled"]
        else:
            reservations = [r for r in all_matching if r.status != "cancelled"]
        wb = Workbook()
        ws = wb.active
        ws.title = "الحجوزات"
        headers = [c[0] for c in RESERVATION_EXCEL_COLUMNS]
        ws.append(headers)
        header_fill = PatternFill(start_color="3FB23F", end_color="3FB23F", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        for r in reservations:
            row = []
            for _, key in RESERVATION_EXCEL_COLUMNS:
                if key == "hotel_name": row.append(r.hotel.name if r.hotel else "")
                elif key == "employee_name": row.append(r.employee.name if r.employee else "")
                elif key == "travel_agent_name": row.append(r.travel_agent.name if r.travel_agent else "")
                elif key == "sales_rep_name": row.append(r.sales_rep.name if r.sales_rep else "")
                elif key in ("created_date", "checkin_date", "checkout_date"):
                    val = getattr(r, key)
                    row.append(val.isoformat() if val else "")
                else: row.append(getattr(r, key, ""))
            ws.append(row)
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = max(12, min(30, max_len + 3))
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return StreamingResponse(
            buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=reservations_export.xlsx"}
        )

    @app.post("/reservations/import/excel")
    async def import_reservations_excel(file: UploadFile = File(...), db: Session = Depends(get_db),
                                        user=Depends(auth.require_permission("reservations.add"))):
        import openpyxl
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return RedirectResponse("/reservations?error=الملف+فاضي", status_code=303)
        header_row = [str(h).strip() if h else "" for h in rows[0]]
        col_index = {}
        for display_name, key in RESERVATION_EXCEL_COLUMNS:
            if display_name in header_row:
                col_index[key] = header_row.index(display_name)
        created, updated = 0, 0
        for row in rows[1:]:
            if not row or all(c is None for c in row): continue
            def get(key, default=""):
                idx = col_index.get(key)
                if idx is None or idx >= len(row): return default
                val = row[idx]
                return val if val is not None else default
            booking_number = str(get("booking_number") or "").strip()
            reservation = db.query(Reservation).filter(Reservation.booking_number == booking_number).first() if booking_number else None
            is_new = reservation is None
            if is_new:
                reservation = Reservation(booking_number=booking_number or generate_booking_number(db),
                                          created_date=_parse_excel_date(get("created_date")) or datetime.date.today())
            reservation.guest_name = str(get("guest_name") or "بدون اسم")
            reservation.checkin_date = _parse_excel_date(get("checkin_date"))
            reservation.checkout_date = _parse_excel_date(get("checkout_date"))
            reservation.adults = int(get("adults") or 1)
            reservation.children = int(get("children") or 0)
            reservation.children_ages = str(get("children_ages") or "")
            reservation.nationality = str(get("nationality") or "")
            reservation.passport_no = str(get("passport_no") or "")
            reservation.phone = str(get("phone") or "")
            hotel = _get_or_create_hotel(db, get("hotel_name"))
            reservation.hotel_id = hotel.id if hotel else reservation.hotel_id
            reservation.room_type = str(get("room_type") or "")
            reservation.stay_type = str(get("stay_type") or "")
            reservation.room_count = int(get("room_count") or 1)
            reservation.company_cost = D(get("company_cost"))
            reservation.stay_cost = D(get("stay_cost"))
            reservation.transportation_cost = D(get("transportation_cost"))
            reservation.other_services_cost = D(get("other_services_cost"))
            rep = _get_or_create_employee(db, get("employee_name"))
            reservation.employee_id = rep.id if rep else None
            reservation.reservation_rep_commission_type = str(get("reservation_rep_commission_type") or "percentage")
            reservation.reservation_rep_commission_value = D(get("reservation_rep_commission_value"))
            agent = _get_or_create_employee(db, get("travel_agent_name"))
            reservation.travel_agent_id = agent.id if agent else None
            reservation.travel_agent_commission_type = str(get("travel_agent_commission_type") or "percentage")
            reservation.travel_agent_commission_value = D(get("travel_agent_commission_value"))
            sales = _get_or_create_employee(db, get("sales_rep_name"))
            reservation.sales_rep_id = sales.id if sales else None
            reservation.sales_rep_commission_type = str(get("sales_rep_commission_type") or "percentage")
            reservation.sales_rep_commission_value = D(get("sales_rep_commission_value"))
            reservation.notes = str(get("notes") or "")
            base = reservation.total_profit
            reservation.employee_commission = Reservation.compute_commission(
                reservation.reservation_rep_commission_type, reservation.reservation_rep_commission_value, base
            ) if reservation.employee_id else DECIMAL_ZERO
            reservation.travel_agent_commission_amount = Reservation.compute_commission(
                reservation.travel_agent_commission_type, reservation.travel_agent_commission_value, base
            ) if reservation.travel_agent_id else DECIMAL_ZERO
            reservation.sales_rep_commission_amount = Reservation.compute_commission(
                reservation.sales_rep_commission_type, reservation.sales_rep_commission_value, base
            ) if reservation.sales_rep_id else DECIMAL_ZERO
            if is_new:
                db.add(reservation)
                created += 1
            else:
                updated += 1
            db.flush()
            from services.commission_service import CommissionService
            CommissionService(db).apply_policies(reservation)
        db.commit()
        return RedirectResponse(f"/reservations?imported=1&created={created}&updated={updated}", status_code=303)

    @app.get("/reservations/{reservation_id}/cancel", response_class=HTMLResponse)
    def cancel_reservation_page(reservation_id: int, request: Request, db: Session = Depends(get_db),
                                 user=Depends(auth.require_permission("reservations.delete"))):
        from services.accounting_service import AccountingService
        r = db.query(Reservation).get(reservation_id)
        if not r:
            return RedirectResponse("/reservations?error=not_found", status_code=303)
        has_accounting = db.query(JournalEntry).filter(
            JournalEntry.source_type == "reservation",
            JournalEntry.source_id == reservation_id,
        ).count() > 0
        return templates.TemplateResponse(request, "reservation_cancel.html", {
            "request": request, "page_title": f"إلغاء الحجز {r.booking_number}",
            "active": "reservations", "r": r,
            "has_accounting": has_accounting,
        })

    @app.post("/reservations/{reservation_id}/cancel")
    def cancel_reservation(reservation_id: int, request: Request, db: Session = Depends(get_db),
                           user=Depends(auth.require_permission("reservations.delete")),
                           reason: str = Form(""), refund_amount: float = Form(0.0),
                           cancellation_fee: float = Form(0.0)):
        from services.reservation_service import ReservationService
        try:
            rsvc = ReservationService(db, user)
            rsvc.cancel(reservation_id, reason=reason,
                        refund_amount=D(refund_amount),
                        cancellation_fee=D(cancellation_fee))
            db.commit()
            return RedirectResponse("/reservations?cancelled=1", status_code=303)
        except ValueError as e:
            return RedirectResponse(f"/reservations/{reservation_id}/cancel?error={str(e)}", status_code=303)

    @app.post("/reservations/{reservation_id}/delete")
    def delete_reservation(reservation_id: int, db: Session = Depends(get_db),
                           user=Depends(auth.require_permission("reservations.delete"))):
        from services.reservation_service import ReservationService
        r = db.query(Reservation).get(reservation_id)
        if not r:
            return RedirectResponse("/reservations?error=not_found", status_code=303)
        if r.status != "draft":
            return RedirectResponse(f"/reservations/{reservation_id}/cancel", status_code=303)
        try:
            ReservationService(db, user).hard_delete(reservation_id)
            db.commit()
            return RedirectResponse("/reservations?deleted=1", status_code=303)
        except ValueError as e:
            return RedirectResponse(f"/reservations?error={str(e)}", status_code=303)

    @app.post("/reservations/{reservation_id}/quick-add-service")
    def quick_add_service(reservation_id: int, name: str = Form(...), price: float = Form(0),
                          db: Session = Depends(get_db),
                          user=Depends(auth.require_permission("reservations.edit"))):
        r = db.query(Reservation).get(reservation_id)
        if not r:
            return RedirectResponse("/reservations", status_code=303)
        svc = Service(name=name, price=D(price))
        db.add(svc)
        db.flush()
        r.services.append(svc)
        db.commit()
        return RedirectResponse(f"/reservations/{reservation_id}/edit?success=added_service", status_code=303)

    @app.post("/reservations/{reservation_id}/remove-service/{service_id}")
    def remove_service(reservation_id: int, service_id: int,
                       db: Session = Depends(get_db),
                       user=Depends(auth.require_permission("reservations.edit"))):
        r = db.query(Reservation).get(reservation_id)
        if r:
            svc = db.query(Service).get(service_id)
            if svc and svc in r.services:
                r.services.remove(svc)
                db.commit()
        return RedirectResponse(f"/reservations/{reservation_id}/edit", status_code=303)

    @app.post("/services/quick-add")
    def quick_add_service_modal(name: str = Form(...), price: float = Form(0),
                                db: Session = Depends(get_db),
                                user=Depends(auth.require_permission("reservations.add"))):
        svc = Service(name=name, price=D(price))
        db.add(svc)
        db.commit()
        return RedirectResponse("/reservations", status_code=303)
