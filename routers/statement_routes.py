import io, datetime
from urllib.parse import quote
from decimal import Decimal
from fastapi import Request, Depends, Form
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from currency_utils import D, DECIMAL_ZERO

from models import (
    Customer, Reservation, Collection, CollectionAllocation,
    SupplierPaymentAllocation,
)
from . import templates, get_db, get_lang
import auth


def _build_statement(db, customer_id, as_of=None):
    """Build full customer statement entries sorted by date.
    Returns (entries: list, totals: dict)."""
    if as_of is None:
        as_of = datetime.date.today()
    c = db.query(Customer).get(customer_id)
    if not c:
        return [], {}

    entries = []

    # 1. Reservations → Debit entries
    reservations = db.query(Reservation).filter(
        Reservation.customer_id == customer_id,
        Reservation.status.notin_(["cancelled", "closed"]),
    ).order_by(Reservation.created_date, Reservation.id).all()

    for r in reservations:
        net = D(r.company_cost or 0) - D(r.discount or 0) + D(r.taxes or 0)
        if net > DECIMAL_ZERO:
            ref_date = r.created_date or r.checkin_date or r.created_date or as_of
            days_overdue = (as_of - ref_date).days if ref_date and ref_date < as_of else 0
            entries.append({
                "date": ref_date,
                "reference": r.booking_number,
                "description": f"حجز #{r.booking_number} - {r.guest_name or ''}",
                "debit": net,
                "credit": DECIMAL_ZERO,
                "balance": DECIMAL_ZERO,
                "days_overdue": max(0, days_overdue),
                "entry_type": "invoice",
                "reservation_id": r.id,
                "is_paid": r.is_paid_in_full,
                "remaining": max(DECIMAL_ZERO, r.remaining_to_office),
            })

    # 2. CollectionAllocations → Credit entries (payments applied to reservations)
    allocations = db.query(CollectionAllocation).join(Collection).filter(
        Collection.customer_id == customer_id,
    ).order_by(CollectionAllocation.date, CollectionAllocation.id).all()

    for a in allocations:
        r = a.reservation
        coll = a.collection
        entries.append({
            "date": a.date or coll.date or as_of,
            "reference": coll.collection_number if coll else "",
            "description": f"تحصيل {coll.collection_number if coll else ''} - {a.amount}",
            "debit": DECIMAL_ZERO,
            "credit": D(a.amount),
            "balance": DECIMAL_ZERO,
            "days_overdue": 0,
            "entry_type": "payment",
            "reservation_id": a.reservation_id,
            "is_paid": True,
            "remaining": DECIMAL_ZERO,
        })

    # 3. Unallocated collection amounts → On-account payments (Credit)
    collections = db.query(Collection).filter(
        Collection.customer_id == customer_id,
        Collection.unallocated_amount > DECIMAL_ZERO,
    ).all()

    for coll in collections:
        if coll.unallocated_amount and D(coll.unallocated_amount) > DECIMAL_ZERO:
            entries.append({
                "date": coll.date or as_of,
                "reference": coll.collection_number,
                "description": f"دفعة تحت الحساب #{coll.collection_number}",
                "debit": DECIMAL_ZERO,
                "credit": D(coll.unallocated_amount),
                "balance": DECIMAL_ZERO,
                "days_overdue": 0,
                "entry_type": "on_account",
                "reservation_id": None,
                "is_paid": True,
                "remaining": DECIMAL_ZERO,
            })

    # Sort by date, then by type (debits first within same date)
    entries.sort(key=lambda e: (e["date"] or as_of, 0 if e["entry_type"] == "invoice" else 1))

    # Calculate running balance
    balance = DECIMAL_ZERO
    for e in entries:
        balance = balance + D(e["debit"]) - D(e["credit"])
        e["balance"] = balance

    # Totals
    totals = {
        "total_debit": sum(D(e["debit"]) for e in entries),
        "total_credit": sum(D(e["credit"]) for e in entries),
        "balance": balance,
        "count": len(entries),
    }

    return entries, totals


def _build_aging(db, customer_id, as_of=None):
    """Build aging report buckets for unpaid invoices."""
    if as_of is None:
        as_of = datetime.date.today()
    buckets = {
        "0_30": {"label": "0-30 يوم", "total": DECIMAL_ZERO, "count": 0, "items": []},
        "31_60": {"label": "31-60 يوم", "total": DECIMAL_ZERO, "count": 0, "items": []},
        "61_90": {"label": "61-90 يوم", "total": DECIMAL_ZERO, "count": 0, "items": []},
        "90_plus": {"label": "أكثر من 90 يوم", "total": DECIMAL_ZERO, "count": 0, "items": []},
    }
    reservations = db.query(Reservation).filter(
        Reservation.customer_id == customer_id,
        Reservation.status.notin_(["cancelled", "closed"]),
    ).all()
    for r in reservations:
        remaining = max(DECIMAL_ZERO, r.remaining_to_office)
        if remaining <= DECIMAL_ZERO:
            continue
        ref_date = r.created_date or r.checkin_date or as_of
        if ref_date:
            days = (as_of - ref_date).days
        else:
            days = 0
        item = {
            "date": ref_date,
            "reference": r.booking_number,
            "description": r.guest_name or "",
            "amount": remaining,
            "days": max(0, days),
            "reservation_id": r.id,
        }
        if days <= 30:
            buckets["0_30"]["items"].append(item)
            buckets["0_30"]["total"] += remaining
            buckets["0_30"]["count"] += 1
        elif days <= 60:
            buckets["31_60"]["items"].append(item)
            buckets["31_60"]["total"] += remaining
            buckets["31_60"]["count"] += 1
        elif days <= 90:
            buckets["61_90"]["items"].append(item)
            buckets["61_90"]["total"] += remaining
            buckets["61_90"]["count"] += 1
        else:
            buckets["90_plus"]["items"].append(item)
            buckets["90_plus"]["total"] += remaining
            buckets["90_plus"]["count"] += 1
    grand_total = sum(b["total"] for b in buckets.values())
    return buckets, grand_total


def setup_statement_routes(app):
    @app.get("/customers/{customer_id}/statement", response_class=HTMLResponse)
    def customer_statement_page(customer_id: int, request: Request, db: Session = Depends(get_db),
                                user=Depends(auth.require_permission("customers.manage"))):
        c = db.query(Customer).get(customer_id)
        if not c:
            return RedirectResponse("/customers", status_code=303)
        entries, totals = _build_statement(db, customer_id)
        aging, aging_total = _build_aging(db, customer_id)
        return templates.TemplateResponse(request, "customer_statement.html", {
            "request": request, "page_title": f"كشف حساب - {c.name}", "active": "customers",
            "customer": c, "entries": entries, "totals": totals,
            "aging": aging, "aging_total": aging_total, "as_of": datetime.date.today(),
        })

    @app.get("/customers/{customer_id}/statement/pdf", response_class=HTMLResponse)
    def customer_statement_pdf(customer_id: int, request: Request, db: Session = Depends(get_db),
                               user=Depends(auth.require_permission("customers.manage"))):
        c = db.query(Customer).get(customer_id)
        if not c:
            return RedirectResponse("/customers", status_code=303)
        entries, totals = _build_statement(db, customer_id)
        aging, aging_total = _build_aging(db, customer_id)
        lang = get_lang(request)
        return templates.TemplateResponse(request, "customer_statement_pdf.html", {
            "request": request, "customer": c, "entries": entries, "totals": totals,
            "aging": aging, "aging_total": aging_total, "as_of": datetime.date.today(),
            "print_mode": True, "lang": lang,
        })

    @app.get("/customers/{customer_id}/statement/excel")
    def customer_statement_excel(customer_id: int, db: Session = Depends(get_db),
                                 user=Depends(auth.require_permission("customers.manage"))):
        c = db.query(Customer).get(customer_id)
        entries, totals = _build_statement(db, customer_id)
        aging, aging_total = _build_aging(db, customer_id)

        wb = Workbook()
        # Sheet 1: Statement
        ws = wb.active
        ws.title = "كشف الحساب"
        header_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin = Side(style="thin", color="D0D0D0")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        ws.merge_cells("A1:F1")
        ws["A1"] = f"كشف حساب - {c.name}"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A2:F2")
        ws["A2"] = f"تاريخ الإصدار: {datetime.date.today().isoformat()}"
        ws["A2"].font = Font(size=10, color="666666")

        headers = ["التاريخ", "المرجع", "الوصف", "مدين", "دائن", "الرصيد"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        for i, e in enumerate(entries, 5):
            ws.cell(row=i, column=1, value=e["date"].isoformat() if e["date"] else "").border = border
            ws.cell(row=i, column=2, value=e["reference"]).border = border
            ws.cell(row=i, column=3, value=e["description"]).border = border
            ws.cell(row=i, column=4, value=float(e["debit"])).border = border
            ws.cell(row=i, column=5, value=float(e["credit"])).border = border
            ws.cell(row=i, column=6, value=float(e["balance"])).border = border

        # Totals row
        tr = 5 + len(entries)
        ws.cell(row=tr, column=3, value="الإجمالي").font = Font(bold=True)
        ws.cell(row=tr, column=4, value=float(totals["total_debit"])).font = Font(bold=True)
        ws.cell(row=tr, column=5, value=float(totals["total_credit"])).font = Font(bold=True)
        ws.cell(row=tr, column=6, value=float(totals["balance"])).font = Font(bold=True, color="CC0000")

        for col_letter in ["A", "B", "C", "D", "E", "F"]:
            ws.column_dimensions[col_letter].width = max(12, 20)

        # Sheet 2: Aging
        ws2 = wb.create_sheet("التقادم الزمني")
        ws2.merge_cells("A1:D1")
        ws2["A1"] = f"تقرير التقادم الزمني - {c.name}"
        ws2["A1"].font = Font(bold=True, size=14)
        ws2.merge_cells("A2:D2")
        ws2["A2"] = f"تاريخ الإصدار: {datetime.date.today().isoformat()}"
        ws2["A2"].font = Font(size=10, color="666666")

        aging_headers = ["الفئة", "العدد", "الإجمالي", ""]
        for col, h in enumerate(aging_headers[:3], 1):
            cell = ws2.cell(row=4, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        ag_row = 5
        for key in ["0_30", "31_60", "61_90", "90_plus"]:
            b = aging[key]
            ws2.cell(row=ag_row, column=1, value=b["label"]).border = border
            ws2.cell(row=ag_row, column=2, value=b["count"]).border = border
            ws2.cell(row=ag_row, column=3, value=float(b["total"])).border = border
            ag_row += 1
        ws2.cell(row=ag_row, column=1, value="الإجمالي").font = Font(bold=True)
        ws2.cell(row=ag_row, column=3, value=float(aging_total)).font = Font(bold=True)

        ws2.column_dimensions["A"].width = 20
        ws2.column_dimensions["B"].width = 12
        ws2.column_dimensions["C"].width = 16

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        safe_name = quote(c.name.replace(" ", "_"))
        return StreamingResponse(
            buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=statement_{safe_name}_{datetime.date.today().isoformat()}.xlsx"}
        )

    @app.get("/customers/{customer_id}/statement/send")
    def customer_statement_send(customer_id: int, request: Request, db: Session = Depends(get_db),
                                user=Depends(auth.require_permission("customers.manage"))):
        c = db.query(Customer).get(customer_id)
        if not c:
            return RedirectResponse("/customers", status_code=303)
        email_to = c.email or ""
        subject = quote(f"كشف حساب - {c.name}")
        body = quote(f"كشف حساب {c.name} حتى تاريخ {datetime.date.today().isoformat()}")
        # Open default email client
        mailto = f"mailto:{email_to}?subject={subject}&body={body}"
        return RedirectResponse(mailto, status_code=303)

    # ─── Reservation Statement ────────────────────────────────────────

    @app.get("/reservations/{reservation_id}/statement", response_class=HTMLResponse)
    def reservation_statement_page(reservation_id: int, request: Request, db: Session = Depends(get_db),
                                   user=Depends(auth.require_permission("reservations.view"))):
        r = db.query(Reservation).get(reservation_id)
        if not r:
            return RedirectResponse("/reservations", status_code=303)
        entries, totals = _build_reservation_statement(db, r)
        return templates.TemplateResponse(request, "reservation_statement.html", {
            "request": request, "page_title": f"كشف حساب حجز {r.booking_number}",
            "active": "board",
            "reservation": r, "entries": entries, "totals": totals,
            "as_of": datetime.date.today(),
        })


def _build_reservation_statement(db, reservation, as_of=None):
    """Build statement entries for a single reservation.
    Returns (entries: list, totals: dict)."""
    if as_of is None:
        as_of = datetime.date.today()
    r = reservation
    entries = []

    # 1. Opening invoice — net sale price (what customer owes)
    net_sale = D(r.company_cost or 0) - D(r.discount or 0) + D(r.taxes or 0)
    if net_sale > DECIMAL_ZERO or True:
        ref_date = r.created_date or r.checkin_date or as_of
        entries.append({
            "date": ref_date,
            "reference": r.booking_number,
            "description": f"فاتورة الحجز #{r.booking_number} - {r.guest_name or ''}",
            "debit": net_sale,
            "credit": DECIMAL_ZERO,
            "balance": DECIMAL_ZERO,
            "entry_type": "invoice",
        })

    # 2. CollectionAllocations → Credits (payments received)
    for ca in (r.collection_allocations or []):
        coll = ca.collection
        entries.append({
            "date": ca.date or (coll.date if coll else as_of),
            "reference": coll.collection_number if coll else "",
            "description": f"تحصيل {coll.collection_number if coll else ''} - {ca.amount}",
            "debit": DECIMAL_ZERO,
            "credit": D(ca.amount),
            "balance": DECIMAL_ZERO,
            "entry_type": "payment",
        })

    # 3. SupplierPaymentAllocations → Debits (payments made for this booking)
    sp_allocations = db.query(SupplierPaymentAllocation).filter(
        SupplierPaymentAllocation.reservation_id == r.id,
    ).all()
    for spa in sp_allocations:
        sp = spa.payment
        entries.append({
            "date": spa.date or (sp.date if sp else as_of),
            "reference": sp.payment_number if sp else "",
            "description": f"دفعة لمورد #{sp.payment_number if sp else ''} - {spa.amount}",
            "debit": D(spa.amount),
            "credit": DECIMAL_ZERO,
            "balance": DECIMAL_ZERO,
            "entry_type": "supplier_payment",
        })

    # 4. If cancelled with refund: show refund credit
    if r.status == "cancelled":
        refund = D(getattr(r, "cancellation_reason_refund", None) or DECIMAL_ZERO)
        if refund > DECIMAL_ZERO:
            entries.append({
                "date": r.cancelled_at.date() if r.cancelled_at else as_of,
                "reference": r.booking_number,
                "description": "مرتجع (إلغاء الحجز)",
                "debit": DECIMAL_ZERO,
                "credit": refund,
                "balance": DECIMAL_ZERO,
                "entry_type": "refund",
            })

    # Sort by date
    entries.sort(key=lambda e: (e["date"] or as_of, 0 if e["entry_type"] == "invoice" else 1))

    # Calculate running balance
    balance = DECIMAL_ZERO
    for e in entries:
        balance = balance + D(e["debit"]) - D(e["credit"])
        e["balance"] = balance

    totals = {
        "total_debit": sum(D(e["debit"]) for e in entries),
        "total_credit": sum(D(e["credit"]) for e in entries),
        "balance": balance,
        "count": len(entries),
        "net_sale": net_sale,
        "total_paid": sum(D(e["credit"]) for e in entries if e["entry_type"] == "payment"),
        "remaining": max(DECIMAL_ZERO, net_sale - sum(D(e["credit"]) for e in entries if e["entry_type"] == "payment")),
    }

    return entries, totals
